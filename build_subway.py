#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""route_transit_feed — 福岡市地下鉄 時刻表フィード（Phase3-A）。

福岡市交通局の公式Excel（空港線・箱崎線 / 七隈線）を自動取得してパースし、
アプリが読む静的JSONを生成する。route_poi_feed と同じ設計思想
（公式源のみ・アプリはスクレイプしない・GitHub Pages配信・秘密情報ゼロ）。

出力: out/subway_timetable.json
  {version, generated_at, source, attribution, stations:{駅:{schedules:[
     {line, direction, day_type(weekday/saturday/holiday), departures:["HH:MM"]}]}}}
  ※アプリ(TrainService)の既存スキーマと一致。

本番はGitHub Actionsで実行→GitHub Pages配信。手動更新(ダイヤ改正時)を不要にする。
パース処理は tools/parse_subway_timetable.py（実績あり）を移植。
"""
from __future__ import annotations
import io
import json
import re
import sys
from datetime import datetime, timezone

import requests
import xlrd
from openpyxl import load_workbook

UA = "route_transit_feed/0.1 (+https://github.com/taxilog-app/route_transit_feed)"
BASE = "https://subway.city.fukuoka.lg.jp/subway/about/data"
URL_KUKO = f"{BASE}/kukohakozaki_timetable.xls"       # 空港線・箱崎線
URL_NANAKUMA = f"{BASE}/nanakuma_timetable.xlsx"       # 七隈線
SOURCE_PAGE = "https://subway.city.fukuoka.lg.jp/subway/about/material.php"

# 公開ガード: 取得崩れで極端に少ない駅数を配信しない（CI失敗→前回分を温存）。
MIN_STATIONS = 30

DAY_MAP = {"平日": "weekday", "土曜": "saturday", "休日": "holiday"}

# 七隈線の同名駅は Excel上「素の名前(橋本方面)」と「(福岡県)(博多方面)」に分裂している。
# アプリは駅名を完全一致で引くため、(福岡県)側を素の名前へ統合し(福岡県)キーを消す。
DUP_PAIRS = [
    ("別府", "別府(福岡県)"), ("梅林", "梅林(福岡県)"), ("橋本", "橋本(福岡県)"),
    ("茶山", "茶山(福岡県)"), ("金山", "金山(福岡県)"),
]


def _download(url: str) -> bytes:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=30)
    r.raise_for_status()
    return r.content


def _excel_time_to_hhmm(v) -> str | None:
    if v is None or v == "":
        return None
    if hasattr(v, "hour") and hasattr(v, "minute"):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, str):
        m = re.match(r"^(\d{1,2}):(\d{2})", v.strip())
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}" if m else None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f < 0 or f >= 1.0:
        return None
    total_min = round(f * 24 * 60)
    return f"{(total_min // 60) % 24:02d}:{total_min % 60:02d}"


def _rows_xls(sheet) -> tuple[list[str], list[list[str | None]]]:
    stations, matrix = [], []
    for r in range(3, sheet.nrows):
        name = str(sheet.cell_value(r, 0)).strip()
        if not name:
            continue
        stations.append(name)
        matrix.append([_excel_time_to_hhmm(sheet.cell_value(r, c))
                       for c in range(2, sheet.ncols)])
    return stations, matrix


def _rows_xlsx(ws) -> tuple[list[str], list[list[str | None]]]:
    rows = list(ws.iter_rows(values_only=True))
    stations, matrix = [], []
    for r in range(3, len(rows)):
        row = rows[r]
        if not row or row[0] in (None, ""):
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        stations.append(name)
        matrix.append([_excel_time_to_hhmm(row[c]) for c in range(2, len(row))])
    return stations, matrix


def _aggregate(by_station, line, direction, day_type, stations, matrix):
    if not matrix:
        return
    n_trains = len(matrix[0])
    for i, st in enumerate(stations):
        deps = [matrix[i][t] for t in range(n_trains) if matrix[i][t]]
        if not deps:
            continue
        deps = sorted(set(deps))
        entry = by_station.setdefault(st, {"schedules": []})
        for sc in entry["schedules"]:
            if sc["line"] == line and sc["direction"] == direction \
                    and sc["day_type"] == day_type:
                sc["departures"] = sorted(set(sc["departures"] + deps))
                break
        else:
            entry["schedules"].append({
                "line": line, "direction": direction,
                "day_type": day_type, "departures": deps,
            })


def _sheet_meta(name: str):
    """シート名 '(平日　姪浜方面)' → (day_type, direction)。該当外は None。"""
    s = name.strip("()").replace("　", " ").strip()
    parts = s.split(" ", 1)
    if len(parts) != 2:
        return None
    day_type = DAY_MAP.get(parts[0].strip())
    return (day_type, parts[1].strip()) if day_type else None


def parse_kuko(content: bytes) -> dict:
    book = xlrd.open_workbook(file_contents=content)
    result = {}
    for sn in book.sheet_names():
        meta = _sheet_meta(sn)
        if not meta:
            continue
        day_type, direction = meta
        stations, matrix = _rows_xls(book.sheet_by_name(sn))
        _aggregate(result, "空港線・箱崎線", direction, day_type, stations, matrix)
    return result


def parse_nanakuma(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    result = {}
    for sn in wb.sheetnames:
        meta = _sheet_meta(sn)
        if not meta:
            continue
        day_type, direction = meta
        stations, matrix = _rows_xlsx(wb[sn])
        _aggregate(result, "七隈線", direction, day_type, stations, matrix)
    return result


def merge(*dicts) -> dict:
    out = {}
    for d in dicts:
        for st, val in d.items():
            if st not in out:
                out[st] = val
            else:
                out[st]["schedules"].extend(val["schedules"])
    return out


def normalize(stations: dict) -> dict:
    """アプリが使う形に正規化。B)同名駅の方向分裂を統合 A)駅座標を付与。
    座標は同梱の subway_station_coords.json（地下鉄駅は不動＝安定）から。"""
    # B) (福岡県)側を素の名前へ統合
    for base, dup in DUP_PAIRS:
        if dup not in stations:
            continue
        base_sc = stations.setdefault(base, {"schedules": []}).setdefault("schedules", [])
        seen = {(s.get("line"), s.get("direction"), s.get("day_type")) for s in base_sc}
        for s in stations[dup].get("schedules", []):
            k = (s.get("line"), s.get("direction"), s.get("day_type"))
            if k not in seen:
                base_sc.append(s)
                seen.add(k)
        del stations[dup]
    # A) 座標付与
    import os
    coord_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "subway_station_coords.json")
    with open(coord_path, encoding="utf-8") as f:
        coords = json.load(f)
    missing = []
    for name, v in stations.items():
        c = coords.get(name)
        if c:
            v["lat"], v["lng"] = c["lat"], c["lng"]
        else:
            missing.append(name)
    if missing:
        print(f"[座標なし駅(要手当)]: {missing}", file=sys.stderr)
    return stations


def main():
    print(f"📥 {URL_KUKO}")
    kuko = parse_kuko(_download(URL_KUKO))
    print(f"  → {len(kuko)}駅")
    print(f"📥 {URL_NANAKUMA}")
    nana = parse_nanakuma(_download(URL_NANAKUMA))
    print(f"  → {len(nana)}駅")
    stations = normalize(merge(kuko, nana))
    print(f"✅ 合計 {len(stations)}駅（正規化後）")

    if len(stations) < MIN_STATIONS:
        print(f"[GUARD] {len(stations)}駅 < 下限{MIN_STATIONS} → 配信中止(異常)",
              file=sys.stderr)
        sys.exit(1)

    import os
    os.makedirs("out", exist_ok=True)
    data = {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": SOURCE_PAGE,
        "attribution": "出典: 福岡市交通局（福岡市地下鉄 時刻表）",
        "stations": stations,
    }
    with open("out/subway_timetable.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    kb = os.path.getsize("out/subway_timetable.json") / 1024
    print(f"📝 out/subway_timetable.json ({kb:.1f} KB)")


if __name__ == "__main__":
    main()
